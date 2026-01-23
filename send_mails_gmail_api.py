import os
import base64
import time
import csv
import shutil
from datetime import datetime
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google.auth.exceptions import RefreshError
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel reports will not be generated.")
    print("Install it with: pip install openpyxl")

# ---------------------------------------------------
# Gmail API scope: only send emails
SCOPES = ["https://www.googleapis.com/auth/gmail.send"]
# --------- CONFIGURE THESE ---------
SUBJECT = "Application for Software Engineer / Backend / MLOps Role"

BODY = """Hi Team,

I hope you are doing well. I am writing to express my interest in Software Engineer / Backend / MLOps roles at your organization. I have attached my resume for your review.

I am currently working as a Software Engineer at OneARVO Ventures, where I lead development of production-grade systems involving Python (FastAPI), Flutter, computer vision, MLOps, and cloud-native deployments. My work spans building scalable backend services, mobile applications, and end-to-end CI/CD pipelines, and deploying ML systems on AWS and GCP.

A few highlights from my experience:
• Led a small Android/Flutter team and delivered multiple applications from development to stable production.
• Built FastAPI-based microservices for computer-vision authentication, achieving >95% accuracy and reducing inference latency by ~40%.
• Designed CI/CD pipelines using GitHub Actions with Docker for automated build and cloud deployment.
• Deployed ML and backend services on AWS (SageMaker, EC2, S3) and GCP (Cloud Run).
• Implemented MLOps workflows using MLflow and Airflow for automated training, versioning, and rollout.

I enjoy working on real-world engineering problems, taking systems from development to reliable production, and building scalable cloud-native solutions. I would welcome the opportunity to contribute to your team and learn from experienced engineers.

Thank you for your time and consideration. I look forward to hearing from you.

Warm regards,  
Aviral Sharma  
+91 9355319465  
aviral.31@outlook.com  

LinkedIn: https://linkedin.com/in/aviral31  
GitHub: https://github.com/aviralsharmaa
Portfolio: https://aviral.info
"""

RESUME_PATH = "Aviral_cv3.pdf"
RECIPIENTS_FILE = r"V:\Aviral\job-mailer\emails_from_excel.txt"
SENT_MAIL_FILE = "sent_mail.txt"           # Only successfully sent emails
FAILED_MAIL_FILE = "not_sent.txt"         # Failed/undelivered emails

BATCH_SIZE = 20                            # Emails per batch
SLEEP_BETWEEN_BATCHES = 60                 # Seconds between batches
RATE_LIMIT_RETRY_DELAY = 300              # 5 minutes wait when rate limited
MAX_EMAILS = 5                          # Maximum emails to send per run

LOG_FILE = "mail_log.csv"                  # CSV log file
EXCEL_REPORT_FILE = "sent_mail_report.xlsx" # Excel report file
# ---------------------------------------------------


def get_gmail_service():
    """Authorize and return a Gmail API service."""
    creds = None
    # token.json stores the user's access and refresh tokens
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)

    # If there are no (valid) credentials, let user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except RefreshError:
                # Token has been expired or revoked, delete it and re-authenticate
                print("⚠️  Token expired or revoked. Re-authenticating...")
                if os.path.exists("token.json"):
                    os.remove("token.json")
                flow = InstalledAppFlow.from_client_secrets_file(
                    "credentials.json", SCOPES
                )
                creds = flow.run_local_server(port=8080)
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json", SCOPES
            )
            creds = flow.run_local_server(port=8080)
        # Save the credentials for next runs
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    service = build("gmail", "v1", credentials=creds)
    return service


def load_recipients(path: str):
    """Load emails from file."""
    with open(path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def init_log_file():
    """Initialize log file with headers."""
    log_path = Path(LOG_FILE)
    if not log_path.exists():
        with log_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["timestamp", "email", "status", "error_message", "gmail_message_id"])


def log_result(email: str, status: str, error_message: str = "", message_id: str = ""):
    """Log email sending result to CSV file."""
    timestamp = datetime.now().isoformat(timespec="seconds")
    with Path(LOG_FILE).open("a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([timestamp, email, status, error_message, message_id])


def move_to_sent_mail(email: str):
    """Move an email to the sent_mail.txt file after successful send."""
    sent_path = Path(SENT_MAIL_FILE)
    # Read existing emails to avoid duplicates
    existing_emails = set()
    if sent_path.exists():
        with sent_path.open("r", encoding="utf-8") as f:
            existing_emails = {line.strip() for line in f if line.strip()}
    
    # Only add if not already present
    if email not in existing_emails:
        with sent_path.open("a", encoding="utf-8") as f:
            f.write(email + "\n")


def move_to_failed_mail(email: str, error_reason: str = ""):
    """Move an email to the not_sent.txt file after delivery failure."""
    failed_path = Path(FAILED_MAIL_FILE)
    # Read existing emails to avoid duplicates
    existing_emails = set()
    if failed_path.exists():
        with failed_path.open("r", encoding="utf-8") as f:
            existing_emails = {line.strip().split(" | ")[0] for line in f if line.strip()}
    
    # Only add if not already present
    if email not in existing_emails:
        with failed_path.open("a", encoding="utf-8") as f:
            if error_reason:
                f.write(f"{email} | {error_reason}\n")
            else:
                f.write(f"{email}\n")


def is_rate_limit_error(error_message: str) -> bool:
    """Check if the error is a rate limit error (429)."""
    error_lower = str(error_message).lower()
    rate_limit_patterns = [
        "429",
        "httperror 429",
        "rate limit",
        "quota exceeded",
        "too many requests",
        "user rate limit exceeded",
    ]
    return any(pattern in error_lower for pattern in rate_limit_patterns)


def is_delivery_failure_error(error_message: str) -> bool:
    """Check if the error indicates delivery failure (invalid address, not found, etc.)."""
    error_lower = str(error_message).lower()
    delivery_failure_patterns = [
        "address not found",
        "mailbox does not exist",
        "invalid email address",
        "invalid recipient",
        "recipient address rejected",
        "user unknown",
        "no such user",
        "mailbox unavailable",
        "550",
        "551",
        "553",
        "invalid recipient address",
        "delivery has failed",
        "delivery status notification",
        "mail delivery subsystem",
        "permanent failure",
        "invalid recipient",
    ]
    return any(pattern in error_lower for pattern in delivery_failure_patterns)


def create_message(sender: str, to: str, subject: str, body_text: str, attachment_path: str):
    """Create a MIME message with optional PDF attachment."""
    message = MIMEMultipart()
    message["To"] = to
    message["From"] = sender
    message["Subject"] = subject

    # Body
    message.attach(MIMEText(body_text, "plain"))

    # Attachment
    resume_file = Path(attachment_path)
    if resume_file.exists():
        with resume_file.open("rb") as f:
            part = MIMEBase("application", "pdf")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{resume_file.name}"',
        )
        message.attach(part)
    else:
        print(f"⚠️ Resume file not found: {attachment_path} (sending without attachment)")

    # Encode to base64 for Gmail API
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    return {"raw": raw_message}


def send_message(service, user_id: str, message_body: dict):
    """Send email using Gmail API."""
    message = service.users().messages().send(userId=user_id, body=message_body).execute()
    return message  # contains id, threadId, etc.


def create_excel_report(sent_emails_data: list):
    """Create or append to Excel report with sent email details."""
    if not EXCEL_AVAILABLE:
        print("⚠️  openpyxl not available. Skipping Excel report generation.")
        return
    
    print(f"\n📊 Creating/Updating Excel report: {EXCEL_REPORT_FILE}...")
    
    # Headers
    headers = ["Timestamp", "Email Address", "Status", "Gmail Message ID", "Error Message"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Try to load existing workbook, or create new one
    report_path = Path(EXCEL_REPORT_FILE)
    if report_path.exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_REPORT_FILE)
            ws = wb.active
            start_row = ws.max_row + 1  # Append after existing data
        except Exception as e:
            print(f"  ⚠️  Could not load existing report: {e}. Creating new one.")
            wb = Workbook()
            ws = wb.active
            ws.title = "Sent Emails Report"
            start_row = 2
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sent Emails Report"
        start_row = 2
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    sent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    
    for row_idx, email_data in enumerate(sent_emails_data, start=start_row):
        timestamp, email, status, message_id, error_message = email_data
        
        ws.cell(row=row_idx, column=1, value=timestamp)
        ws.cell(row=row_idx, column=2, value=email)
        ws.cell(row=row_idx, column=3, value=status)
        ws.cell(row=row_idx, column=4, value=message_id)
        ws.cell(row=row_idx, column=5, value=error_message)
        
        # Color code rows based on status
        row_fill = sent_fill if status == "SENT" else failed_fill
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = row_fill
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20  # Timestamp
    ws.column_dimensions['B'].width = 35  # Email
    ws.column_dimensions['C'].width = 15  # Status
    ws.column_dimensions['D'].width = 30  # Message ID
    ws.column_dimensions['E'].width = 50  # Error Message
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook with retry logic for permission errors
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            # Try to save directly
            wb.save(EXCEL_REPORT_FILE)
            print(f"✅ Excel report saved: {EXCEL_REPORT_FILE} ({len(sent_emails_data)} new entries)")
            return
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"  ⚠️  File is locked (may be open in Excel). Retrying in {retry_delay} seconds... (Attempt {attempt + 1}/{max_retries})")
                time.sleep(retry_delay)
            else:
                # Last attempt failed, try saving to a temporary file
                try:
                    temp_file = f"{EXCEL_REPORT_FILE}.tmp"
                    wb.save(temp_file)
                    # Try to rename (might still fail if original is locked)
                    try:
                        if Path(EXCEL_REPORT_FILE).exists():
                            Path(EXCEL_REPORT_FILE).unlink()
                        shutil.move(temp_file, EXCEL_REPORT_FILE)
                        print(f"✅ Excel report saved: {EXCEL_REPORT_FILE} ({len(sent_emails_data)} new entries)")
                    except Exception:
                        # Keep temp file if rename fails
                        print(f"  ⚠️  Could not overwrite {EXCEL_REPORT_FILE} (file may be open).")
                        print(f"  ✅ Saved to temporary file: {temp_file}")
                        print(f"  💡 Please close {EXCEL_REPORT_FILE} and rename {temp_file} manually.")
                except Exception as e:
                    print(f"  ❌ Could not save Excel report: {e}")
                    print(f"  💡 Please close {EXCEL_REPORT_FILE} if it's open and try again.")
        except Exception as e:
            print(f"  ❌ Error saving Excel report: {e}")
            return


def move_sent_emails_to_file(sent_emails: set):
    """Move sent emails from emails_from_excel.txt to sent_mail.txt"""
    if not sent_emails:
        return
    
    print(f"\n📝 Moving {len(sent_emails)} sent emails to {SENT_MAIL_FILE}...")
    
    # Read existing sent emails
    existing_sent = set()
    sent_path = Path(SENT_MAIL_FILE)
    if sent_path.exists():
        with sent_path.open("r", encoding="utf-8") as f:
            existing_sent = {line.strip() for line in f if line.strip()}
    
    # Add new sent emails
    all_sent_emails = existing_sent | sent_emails
    
    # Write all sent emails to sent_mail.txt
    with sent_path.open("w", encoding="utf-8") as f:
        for email in sorted(all_sent_emails):
            f.write(email + "\n")
    
    print(f"✅ Total emails in {SENT_MAIL_FILE}: {len(all_sent_emails)}")


def move_failed_emails_to_file(failed_emails: dict):
    """Move failed emails to not_sent.txt"""
    if not failed_emails:
        return
    
    print(f"\n📝 Moving {len(failed_emails)} failed emails to {FAILED_MAIL_FILE}...")
    
    # Read existing failed emails
    existing_failed = {}
    failed_path = Path(FAILED_MAIL_FILE)
    if failed_path.exists():
        with failed_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split(" | ", 1)
                    email = parts[0]
                    reason = parts[1] if len(parts) > 1 else ""
                    existing_failed[email] = reason
    
    # Merge with new failed emails (new errors take precedence)
    all_failed_emails = {**existing_failed, **failed_emails}
    
    # Write all failed emails to not_sent.txt
    with failed_path.open("w", encoding="utf-8") as f:
        for email in sorted(all_failed_emails.keys()):
            reason = all_failed_emails[email]
            if reason:
                f.write(f"{email} | {reason}\n")
            else:
                f.write(f"{email}\n")
    
    print(f"✅ Total failed emails in {FAILED_MAIL_FILE}: {len(all_failed_emails)}")


def main():
    init_log_file()
    recipients = load_recipients(RECIPIENTS_FILE)
    print(f"Total recipients: {len(recipients)}")

    service = get_gmail_service()
    user_id = "me"

    sent_count = 0
    failed_count = 0
    delivery_failed_count = 0
    batch_count = 0
    sent_emails = set()  # Track successfully sent emails (only truly sent)
    failed_emails = {}  # Track failed emails with error reasons
    sent_emails_data = []  # Track sent email details for Excel report

    for i, email in enumerate(recipients, start=1):
        try:
            mime_message = create_message(
                sender="",   # Gmail API uses authenticated user; "me" is enough
                to=email,
                subject=SUBJECT,
                body_text=BODY,
                attachment_path=RESUME_PATH,
            )

            # Send email
            result = send_message(service, user_id, mime_message)
            message_id = result.get("id", "")
            status = "SENT"
            error_message = ""

            sent_count += 1
            batch_count += 1
            sent_emails.add(email)  # Track successful send
            
            # Track for Excel report
            timestamp = datetime.now().isoformat(timespec="seconds")
            sent_emails_data.append([timestamp, email, status, message_id, error_message])
            
            # Move to sent_mail.txt immediately after successful send
            move_to_sent_mail(email)
            
            print(f"[{i}/{len(recipients)}] ✅ Sent to {email} (message id: {message_id})")
            
            # Check if we've reached the maximum email limit
            if sent_count >= MAX_EMAILS:
                print(f"\n⏹️  Reached maximum email limit ({MAX_EMAILS}). Stopping...")
                break

        except Exception as e:
            error_message = str(e)
            message_id = ""
            
            # Check if this is a delivery failure (invalid address, not found, etc.)
            if is_delivery_failure_error(error_message):
                status = "DELIVERY_FAILED"
                delivery_failed_count += 1
                failed_emails[email] = error_message[:100]  # Store error reason
                print(f"[{i}/{len(recipients)}] ❌ Delivery failed: {email} - {error_message[:100]}")
                # Move to failed file immediately
                move_to_failed_mail(email, error_message[:100])
            # Check if this is a rate limit error (temporary, should retry)
            elif is_rate_limit_error(error_message):
                status = "RATE_LIMITED"
                failed_count += 1
                print(f"[{i}/{len(recipients)}] ⚠️  Rate limited: {email}")
                print(f"    Waiting {RATE_LIMIT_RETRY_DELAY // 60} minutes before continuing...")
                time.sleep(RATE_LIMIT_RETRY_DELAY)  # Wait longer for rate limit
                batch_count = 0  # Reset batch counter after rate limit wait
            else:
                # Other errors (network, etc.) - might be temporary
                status = "FAILED"
                failed_count += 1
                print(f"[{i}/{len(recipients)}] ❌ Failed to send to {email}: {error_message[:100]}")
            
            # Track failed emails for report
            timestamp = datetime.now().isoformat(timespec="seconds")
            sent_emails_data.append([timestamp, email, status, message_id, error_message[:200]])

        # Log every attempt
        log_result(email, status, error_message, message_id)

        # Batching to be safe with quotas
        if batch_count >= BATCH_SIZE:
            print(f"⏸️ Pausing {SLEEP_BETWEEN_BATCHES} seconds to avoid limits...")
            time.sleep(SLEEP_BETWEEN_BATCHES)
            batch_count = 0

    # Load existing sent emails from sent_mail.txt
    existing_sent_emails = set()
    sent_path = Path(SENT_MAIL_FILE)
    if sent_path.exists():
        with sent_path.open("r", encoding="utf-8") as f:
            existing_sent_emails = {line.strip() for line in f if line.strip()}
    
    # Load existing failed emails from not_sent.txt (to exclude from retries)
    existing_failed_emails = set()
    failed_path = Path(FAILED_MAIL_FILE)
    if failed_path.exists():
        with failed_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    # Extract email (before " | " if present)
                    email = line.split(" | ")[0]
                    existing_failed_emails.add(email)
    
    # Combine all sent emails (sent in this run + already in sent_mail.txt)
    all_sent_emails = sent_emails | existing_sent_emails
    
    # Combine all failed emails (failed in this run + already in not_sent.txt)
    all_failed_emails = set(failed_emails.keys()) | existing_failed_emails
    
    # Keep only emails that haven't been sent AND haven't failed permanently
    remaining_emails = [
        email for email in recipients 
        if email not in all_sent_emails and email not in all_failed_emails
    ]
    
    # Update recipients file (emails_from_excel.txt) with remaining emails
    if remaining_emails:
        print(f"\n📝 Updating {RECIPIENTS_FILE} with remaining {len(remaining_emails)} emails...")
        with Path(RECIPIENTS_FILE).open("w", encoding="utf-8") as f:
            for email in remaining_emails:
                f.write(email + "\n")
    else:
        print(f"\n📝 All emails processed!")
    
    # Move sent emails to sent_mail.txt (consolidate all sent emails)
    if sent_emails:
        move_sent_emails_to_file(sent_emails)
    
    # Move failed emails to not_sent.txt
    if failed_emails:
        move_failed_emails_to_file(failed_emails)
    
    # Create Excel report with all sent email details
    if sent_emails_data:
        create_excel_report(sent_emails_data)

    print("\n" + "="*60)
    print("✅ Finished.")
    print("="*60)
    print(f"Total recipients: {len(recipients)}")
    print(f"✅ Successfully sent: {sent_count} (limit: {MAX_EMAILS})")
    print(f"❌ Delivery failed (invalid addresses): {delivery_failed_count}")
    print(f"⚠️  Failed (temporary errors - will retry): {failed_count}")
    if failed_count > 0:
        print(f"   Note: Rate limit errors (429) are temporary and will be retried on next run")
    print(f"📋 Remaining in {RECIPIENTS_FILE}: {len(remaining_emails)}")
    print(f"✅ Sent emails saved to: {SENT_MAIL_FILE} (total: {len(all_sent_emails)})")
    print(f"❌ Failed emails saved to: {FAILED_MAIL_FILE} (total: {len(failed_emails)})")
    print(f"📊 CSV Log saved to: {LOG_FILE}")
    if EXCEL_AVAILABLE and sent_emails_data:
        print(f"📊 Excel Report saved to: {EXCEL_REPORT_FILE}")
    print("="*60)


if __name__ == "__main__":
    main()
