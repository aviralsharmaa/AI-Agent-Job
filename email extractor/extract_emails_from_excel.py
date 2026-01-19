import re
import time
import openpyxl
from pathlib import Path
from typing import List, Set, Optional


def extract_emails_from_text(text: str) -> List[str]:
    """Extract all email addresses from a text string"""
    if not text or not isinstance(text, str):
        return []
    
    # Regex pattern for email addresses (more comprehensive)
    email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
    emails = re.findall(email_pattern, str(text))
    
    # Clean and normalize emails
    cleaned = []
    seen_in_text = set()
    for email in emails:
        email = email.strip().lower()
        if email and email not in seen_in_text:
            cleaned.append(email)
            seen_in_text.add(email)
    
    return cleaned


def is_email_column(header_text: str) -> bool:
    """Check if a column header indicates it contains email addresses"""
    if not header_text:
        return False
    
    header_lower = str(header_text).lower().strip()
    email_keywords = [
        'email', 'mail', 'e-mail', 'e_mail', 'mailid', 'mail id',
        'email address', 'emailaddress', 'email_addr', 'email id',
        'contact email', 'contactemail', 'correo', 'courriel'
    ]
    
    return any(keyword in header_lower for keyword in email_keywords)


def extract_emails_from_excel(excel_file: str, output_file: str = None) -> List[str]:
    """
    Extract all email addresses from an Excel file.
    Optimized for large datasets (30K+ rows).
    Scans all cells but prioritizes columns with email-related headers.
    
    Args:
        excel_file: Path to the Excel file
        output_file: Optional path to save emails (default: emails_from_excel.txt)
    
    Returns:
        List of unique email addresses
    """
    print(f"Reading Excel file: {excel_file}...")
    print("This may take a while for large files (30K+ rows)...\n")
    
    # Load the workbook (try read_only mode first for better performance with large files)
    wb = None
    try:
        print("  Attempting to load in read-only mode (faster for large files)...")
        wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
    except FileNotFoundError:
        print(f"❌ Error: File '{excel_file}' not found!")
        return []
    except Exception as e:
        # Fallback to regular mode if read_only fails
        print(f"  Read-only mode failed: {e}")
        print("  Falling back to regular mode...")
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e2:
            print(f"❌ Error loading Excel file: {e2}")
            return []
    
    all_emails: Set[str] = set()
    total_rows = 0
    email_columns = []
    start_time = time.time()  # Track total processing time
    
    # Process all sheets
    for sheet_name in wb.sheetnames:
        print(f"\nProcessing sheet: '{sheet_name}'...")
        ws = wb[sheet_name]
        
        # Find email columns by checking header row
        header_row = None
        email_columns = []
        
        # Try to read header row (handle read-only mode where max_row might be None)
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                email_columns = [
                    idx for idx, header in enumerate(header_row)
                    if is_email_column(str(header) if header else "")
                ]
                if email_columns:
                    print(f"  ✓ Found {len(email_columns)} email column(s) by header name")
                else:
                    print(f"  ℹ No email columns found by header - scanning all columns")
        except Exception as e:
            print(f"  ℹ Could not read header row - scanning all columns")
        
        # Process rows efficiently
        start_row = 2 if header_row else 1  # Skip header if found
        sheet_rows = 0
        sheet_start_time = time.time()  # Track time for this sheet
        
        print(f"  Starting extraction...")
        
        try:
            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                total_rows += 1
                sheet_rows += 1
                
                # Show real-time progress every 1000 rows (more frequent updates)
                if total_rows % 1000 == 0:
                    elapsed = time.time() - start_time  # Use total start_time for cumulative progress
                    rate = total_rows / elapsed if elapsed > 0 else 0
                    elapsed_str = f"{elapsed:.1f}s" if elapsed < 60 else f"{elapsed/60:.1f}m"
                    print(f"  ⏳ Processed {total_rows:,} rows | Found {len(all_emails):,} unique emails | Time: {elapsed_str} | Speed: {rate:.0f} rows/sec")
                
                # Scan all cells for emails
                for col_idx, cell_value in enumerate(row):
                    if cell_value:
                        emails = extract_emails_from_text(str(cell_value))
                        all_emails.update(emails)
        except StopIteration:
            # End of sheet
            pass
        
        sheet_elapsed_time = time.time() - sheet_start_time
        print(f"  ✓ Completed sheet '{sheet_name}': {sheet_rows:,} rows processed in {sheet_elapsed_time:.1f}s")
    
    # Convert to sorted list
    total_time = time.time() - start_time
    unique_emails = sorted(list(all_emails))
    
    print(f"\n{'='*60}")
    print(f"✅ Extraction Complete!")
    print(f"Total rows processed: {total_rows:,}")
    print(f"Unique emails found: {len(unique_emails):,}")
    print(f"Total time: {total_time:.1f} seconds ({total_time/60:.1f} minutes)")
    if total_rows > 0:
        print(f"Average speed: {total_rows/total_time:.0f} rows/second")
    print(f"{'='*60}")
    
    # Save to file if output_file specified
    if output_file:
        save_emails_to_file(unique_emails, output_file)
    else:
        # Default output file name
        excel_path = Path(excel_file)
        default_output = excel_path.parent / "emails_from_excel.txt"
        save_emails_to_file(unique_emails, str(default_output))
    
    return unique_emails


def save_emails_to_file(emails: List[str], output_file: str):
    """Save emails to a text file, one per line (optimized for large files)"""
    print(f"\nSaving {len(emails):,} emails to: {output_file}...")
    
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            for idx, email in enumerate(emails, 1):
                f.write(email + "\n")
                # Show progress every 10000 emails
                if idx % 10000 == 0:
                    print(f"  Saved {idx:,}/{len(emails):,} emails...")
        
        print(f"✅ Successfully saved {len(emails):,} unique emails to {output_file}")
    except Exception as e:
        print(f"❌ Error saving file: {e}")
        raise


if __name__ == "__main__":
    # Configuration
    excel_file = "23800+ Ultimate HR Outreach List - DataNiti.xlsx"  # Path to your Excel file
    output_file = "emails_from_excel.txt"  # Output file name (all emails, one per line)
    
    print("="*60)
    print("Excel Email Extractor")
    print("Extracts all unique emails from Excel sheets")
    print("="*60)
    print()
    
    # Extract emails
    emails = extract_emails_from_excel(excel_file, output_file)
    
    # Print summary and preview
    if emails:
        print(f"\n📧 Email Extraction Summary:")
        print(f"  Total unique emails: {len(emails):,}")
        print(f"  Output file: {output_file}")
        print(f"\n📧 Preview (first 10 emails):")
        for i, email in enumerate(emails[:10], 1):
            print(f"  {i}. {email}")
        if len(emails) > 10:
            print(f"  ... and {len(emails) - 10:,} more")
        print(f"\n✅ All emails saved to: {output_file}")
    else:
        print("\n⚠️  No emails found in the Excel file.")
