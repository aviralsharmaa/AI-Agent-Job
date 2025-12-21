import re
import openpyxl
from openpyxl import Workbook
from urllib.parse import urlparse
from typing import List, Tuple, Set


def is_url(text: str) -> bool:
    """Check if text is a URL"""
    url_patterns = [
        r'https?://',  # http:// or https://
        r'www\.',      # www.
        r'linkedin\.com',
        r'forms\.gle',
        r'google\.com/forms',
        r'docs\.google\.com',
        r'\.in/',
        r'\.com/',
        r'\.co/',
        r'zurl\.co',
        r'shorturl\.at',
    ]
    text_lower = text.lower().strip()
    for pattern in url_patterns:
        if re.search(pattern, text_lower):
            return True
    return False


def extract_emails_from_line(line: str) -> List[str]:
    """Extract all emails from a line, handling comma-separated emails"""
    # Remove URLs first - comprehensive pattern
    url_pattern = r'https?://[^\s,]+|www\.[^\s,]+|linkedin\.com[^\s,]*|forms\.gle[^\s,]*|docs\.google\.com[^\s,]*|zurl\.co[^\s,]*|shorturl\.at[^\s,]*|\.in/[^\s,]*|\.com/[^\s,]*|\.co/[^\s,]*'
    line = re.sub(url_pattern, '', line, flags=re.IGNORECASE)
    
    # Split by common separators (comma, slash, space) to handle cases like "email1/email2"
    # But be careful not to split valid email domains
    parts = re.split(r'[,/\s]+', line)
    
    # Email pattern
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    emails = []
    
    # Check each part for valid email
    for part in parts:
        part = part.strip().rstrip(',').strip()
        if re.match(email_pattern, part):
            if not is_url(part):
                emails.append(part.lower())  # Normalize to lowercase
    
    # Also try finding emails in the original line (in case splitting broke something)
    email_pattern_find = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    found_emails = re.findall(email_pattern_find, line)
    
    # Merge and deduplicate
    all_emails = list(set(emails + found_emails))
    
    # Final cleanup
    cleaned_emails = []
    for email in all_emails:
        email = email.strip().rstrip(',').strip().lower()
        if email and not is_url(email) and '@' in email:
            # Double-check it's a valid email format
            if re.match(email_pattern, email):
                cleaned_emails.append(email)
    
    return cleaned_emails


def extract_company_name(email: str) -> str:
    """Extract company name from email domain"""
    try:
        domain = email.split('@')[1]
        # Remove common TLDs
        company = domain.split('.')[0]
        # Capitalize first letter
        return company.capitalize()
    except:
        return "Unknown"


def is_software_company(email: str, company_name: str) -> bool:
    """Check if company is software/IT/product-based"""
    # Keywords that indicate software/IT/product companies
    tech_keywords = [
        # Software/IT terms
        'tech', 'software', 'soft', 'systems', 'sys', 'solutions', 'sol',
        'digital', 'data', 'cloud', 'cyber', 'info', 'it', 'ict',
        'dev', 'develop', 'code', 'coder', 'programming', 'program',
        'app', 'apps', 'mobile', 'web', 'internet', 'online',
        'ai', 'ml', 'iot', 'saas', 'paas', 'api', 'sdk',
        'platform', 'product', 'service', 'svc',
        'automation', 'auto', 'robot', 'bot',
        'network', 'net', 'server', 'server',
        'security', 'secure', 'cyber',
        'analytics', 'analysis', 'insight', 'intelligence',
        'consulting', 'consult', 'consultancy',
        'lab', 'labs', 'studio', 'works', 'workshop',
        'innovation', 'innovate', 'venture', 'ventures',
        'startup', 'start', 'scale', 'growth',
        # Specific tech companies/domains
        'microsoft', 'google', 'amazon', 'aws', 'azure',
        'oracle', 'sap', 'salesforce', 'adobe', 'ibm',
        'dell', 'hp', 'cisco', 'intel', 'nvidia',
        'accenture', 'tcs', 'infosys', 'wipro', 'hcl',
        'techmahindra', 'cognizant', 'capgemini', 'mindtree',
        # File extensions that indicate tech
        '.io', '.ai', '.tech', '.app', '.dev', '.cloud',
        # Common tech company suffixes
        'technologies', 'technology', 'techno',
        'informatics', 'informatics', 'informatics',
        'computing', 'computer', 'compute',
        'engineering', 'engineer', 'engg',
    ]
    
    # Check email domain and company name
    email_lower = email.lower()
    company_lower = company_name.lower()
    domain = email_lower.split('@')[1] if '@' in email_lower else ''
    
    # Check for tech keywords in company name
    for keyword in tech_keywords:
        if keyword in company_lower or keyword in domain:
            return True
    
    # Check for common tech TLDs
    tech_tlds = ['.io', '.ai', '.tech', '.app', '.dev', '.cloud']
    for tld in tech_tlds:
        if tld in domain:
            return True
    
    # Check for specific patterns
    tech_patterns = [
        r'\btech\b', r'\bsoft\b', r'\bsys\b', r'\bdev\b',
        r'\bdata\b', r'\bdigital\b', r'\bcloud\b', r'\bai\b',
        r'\bml\b', r'\biot\b', r'\bsaas\b', r'\bapi\b',
    ]
    
    combined_text = f"{company_lower} {domain}"
    for pattern in tech_patterns:
        if re.search(pattern, combined_text):
            return True
    
    return False


def clean_allmail_file(input_file: str) -> Tuple[List[str], List[str]]:
    """Clean allmail.txt and extract emails and company names"""
    all_emails = []
    seen_emails: Set[str] = set()
    
    print(f"Reading {input_file}...")
    with open(input_file, "r", encoding="utf-8") as f:
        lines = f.readlines()
    
    print(f"Processing {len(lines)} lines...")
    
    for line_num, line in enumerate(lines, 1):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Skip if entire line is a URL
        if is_url(line):
            continue
        
        # Extract emails from the line
        emails = extract_emails_from_line(line)
        
        for email in emails:
            # Remove duplicates (case-insensitive)
            email_lower = email.lower()
            if email_lower not in seen_emails:
                seen_emails.add(email_lower)
                all_emails.append(email)
    
    # Extract company names
    company_names = [extract_company_name(email) for email in all_emails]
    
    print(f"Found {len(all_emails)} unique emails")
    return all_emails, company_names


def create_excel(emails: List[str], company_names: List[str], is_tech_list: List[bool], output_file: str):
    """Create Excel file with emails, company names, and tech/non-tech classification"""
    print(f"Creating Excel file: {output_file}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Data"
    
    # Write headers in row 1
    ws['A1'] = "Email Addresses"
    ws['B1'] = "Company Names"
    ws['C1'] = "Company Type"
    
    # Write emails in column A (starting from row 2)
    print("Writing emails to column A...")
    for idx, email in enumerate(emails, start=2):
        ws.cell(row=idx, column=1, value=email)
        if (idx - 1) % 1000 == 0:
            print(f"  Processed {idx - 1} emails...")
    
    # Write company names in column B (starting from row 2)
    print("Writing company names to column B...")
    for idx, company in enumerate(company_names, start=2):
        ws.cell(row=idx, column=2, value=company)
        if (idx - 1) % 1000 == 0:
            print(f"  Processed {idx - 1} companies...")
    
    # Write company type in column C (starting from row 2)
    print("Writing company types to column C...")
    from openpyxl.styles import PatternFill
    
    tech_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    nontech_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink
    
    for idx, is_tech in enumerate(is_tech_list, start=2):
        company_type = "Tech/Software" if is_tech else "Non-Tech"
        cell = ws.cell(row=idx, column=3, value=company_type)
        # Color code: green for tech, pink for non-tech
        cell.fill = tech_fill if is_tech else nontech_fill
        if (idx - 1) % 1000 == 0:
            print(f"  Processed {idx - 1} company types...")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Make header row bold
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font
    ws['C1'].font = bold_font
    
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")


def save_to_input_txt(emails: List[str], output_file: str):
    """Save cleaned emails to input.txt (one per line)"""
    print(f"Saving cleaned emails to {output_file}...")
    
    with open(output_file, "w", encoding="utf-8") as f:
        for email in emails:
            f.write(email + "\n")
    
    print(f"Saved {len(emails)} emails to {output_file}")


def save_tech_emails(emails: List[str], company_names: List[str], is_tech_list: List[bool], output_file: str):
    """Save only software/IT company emails to hireIT.txt"""
    print(f"Saving software/IT company emails to {output_file}...")
    
    tech_emails = [email for email, is_tech in zip(emails, is_tech_list) if is_tech]
    
    with open(output_file, "w", encoding="utf-8") as f:
        for email in tech_emails:
            f.write(email + "\n")
    
    print(f"Saved {len(tech_emails)} software/IT company emails to {output_file}")
    return tech_emails


if __name__ == "__main__":
    input_file = "allmail.txt"
    excel_output = "email_data.xlsx"
    txt_output = "input.txt"
    tech_output = "hireIT.txt"
    
    print("="*60)
    print("Email Manager - Cleaning and Organizing Emails")
    print("="*60)
    
    # Clean and extract emails
    emails, company_names = clean_allmail_file(input_file)
    
    if not emails:
        print("No emails found!")
        exit(1)
    
    # Classify companies as tech/non-tech
    print("\nClassifying companies as Tech/Non-Tech...")
    is_tech_list = []
    tech_count = 0
    nontech_count = 0
    
    for email, company in zip(emails, company_names):
        is_tech = is_software_company(email, company)
        is_tech_list.append(is_tech)
        if is_tech:
            tech_count += 1
        else:
            nontech_count += 1
    
    print(f"Tech/Software companies: {tech_count}")
    print(f"Non-Tech companies: {nontech_count}")
    
    # Create Excel file with tech/non-tech classification
    create_excel(emails, company_names, is_tech_list, excel_output)
    
    # Save all emails to input.txt
    save_to_input_txt(emails, txt_output)
    
    # Save only tech/software company emails to hireIT.txt
    tech_emails = save_tech_emails(emails, company_names, is_tech_list, tech_output)
    
    print("\n" + "="*60)
    print("✔ Processing Complete!")
    print(f"Total unique emails: {len(emails)}")
    print(f"  - Tech/Software companies: {tech_count}")
    print(f"  - Non-Tech companies: {nontech_count}")
    print(f"\nFiles created:")
    print(f"  - {txt_output}: All emails ({len(emails)} emails)")
    print(f"  - {tech_output}: Software/IT companies only ({len(tech_emails)} emails)")
    print(f"  - {excel_output}: Complete data with Tech/Non-Tech classification")
    print("="*60)

